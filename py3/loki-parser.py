# ====================================
# title           : loki-parser.py
# description     : Parse Loki logs to ingest data into an ES instance or a local Excel sheet
# author          : Syed Hasan
# date            : Dec 30 2020
# version         : 0.9
# usage           : python loki-parser.py -h
# tested:         : Windows
# python_version  : 3.X
# ====================================

import re
import argparse
import pprint
import datetime
import time

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from dateutil import parser

from elasticsearch import Elasticsearch
from elasticsearch import helpers

#
# Gllobal Variables
#

reg = "([a-z,A-Z].*:\d\d) ([a-z,A-Z].*) (sshd.*\:) ([A-Z,a-z].*.from)(.\d*.\d*.\d*.\d*)"
reg_keys = ["@timestamp","system","service","message","IP"]

keys_process = ["MODULE:","MESSAGE:","PID:","NAME:","OWNER:","CMD:","PATH:"]
# Deprecated: keys_filescan = ["MODULE:","MESSAGE:","FILE:", "TYPE:","SIZE:","FIRST_BYTES:","MD5:","SHA1:","SHA256:","CREATED:", "MODIFIED:","ACCESSED:","REASON_1:","SCORE:"]
keys_filescan = ["MODULE:","MESSAGE:","FILE:", "SCORE:", "TYPE:","SIZE:","FIRST_BYTES:","MD5:","SHA1:","SHA256:","CREATED:", "MODIFIED:","ACCESSED:","REASON_1:","SUBSCORE:"]

#
# Functions
#

def currDate():
    return datetime.date.fromtimestamp(time.time())

def saveLogs(fileName, data):
    workbook = Workbook()

    fileScanSheet = workbook.create_sheet("File Scan")
    procScanSheet = workbook.create_sheet("Process Scan")
    otherSheet = workbook.create_sheet("Other Logs")

    # Add the timestamp column manually. TODO: Not cool.
    fileScanSheet.cell(row=1, column=1).value = "Timestamp"
    procScanSheet.cell(row=1, column=1).value = "Timestamp"
    otherSheet.cell(row=1, column=1).value = "Timestamp"
    otherSheet.cell(row=1, column=2).value = "Mode"

    # Set the columns in respective sheets
    for count, key in enumerate(keys_filescan, start=2):
        fileScanSheet.cell(row=1, column=count).value = key.rstrip(":").title()

    for count, key in enumerate(keys_process, start=2):
        procScanSheet.cell(row=1, column=count).value = key.rstrip(":").title()

    # Maintain separate row-col numbers. Isn't there a better way to do this?!
    fsRowCount = 2
    fsColCount = 1
    psRowCount = 2
    psColCount = 1
    otRowCount = 2
    otColCount = 1

    # The Real Thing!
    for currCount, value in enumerate(data, start=2):
        if value['MODULE'] == "FileScan":
            for attribute, dictVal in value.items():
                fileScanSheet.cell(row=fsRowCount, column=fsColCount).value = dictVal
                fsColCount += 1

            fsRowCount += 1
            fsColCount = 1

        elif value['MODULE'] == "ProcessScan":
            for attribute, dictVal in value.items():
                procScanSheet.cell(row=psRowCount, column=psColCount).value = dictVal
                psColCount += 1

            psRowCount += 1
            psColCount = 1

        else:
            for attribute, dictVal in value.items():
                otherSheet.cell(row=otRowCount, column=otColCount).value = dictVal
                otColCount += 1

            otRowCount += 1
            otColCount = 1

    # Remove the default sheet. Could've renamed this to the third sheet but...
    sheet = workbook['Sheet']
    workbook.remove(sheet)

    # Style the first row senor
    whiteFont = Font(color='FFFFFF', size='12')
    blackBg = PatternFill(bgColor="000000", fill_type = "solid")

    for cell in procScanSheet["1:1"]:
        cell.font = whiteFont
        cell.fill = blackBg
    for cell in fileScanSheet["1:1"]:
        cell.font = whiteFont
        cell.fill = blackBg
    for cell in otherSheet["1:1"]:
        cell.font = whiteFont
        cell.fill = blackBg

    try:
        workbook.save(filename=f"{fileName}-{currDate()}.xlsx")
        print(f"INFO: Successfully parsed all logs and wrote them to an XLSX sheet. Total Logs: {len(data)}")
    except PermissionError:
        print("ERROR: Please close the workbook before executing write operations on the worksheet.")

def shipLogs(elasticCon, dataBundle):
    helpers.bulk(elasticCon, dataBundle)
    print("INFO: Shipped logs to the Elastic stack")


def parseLogs(filePtr, esConn, mode):
    bundle = []

    try:
        logfileName = re.search("\\w+", filePtr.name).group(0)
    except:
        logfileName = "loki_logfile"

    for line in filePtr.readlines():
        if 'FileScan' == re.search('MODULE:(.*)MESSAGE',line).groups()[0].strip(' '):
            log = {}
            log['@timestamp']= parser.parse(line.split(' ')[0])
            for count,key in enumerate(keys_filescan):
                try:
                    # Old variant: found = re.search(key+'(.*)'+keys_filescan[count+1],line).groups()[0]
                    found = re.search('(?<='+key+')(.*?)(?='+keys_filescan[count+1]+')', line).groups()[0]
                    sanitizedKey = key.rstrip(":")
                    found = found.lstrip(' ').rstrip(' ')
                    log[sanitizedKey]=found
                except:
                    log['SCORE']=line.split(':')[-1].strip()
                    pass

        else:
            log = {}
            log['@timestamp']=parser.parse(line.split(' ')[0])
            for count,key in enumerate(keys_process):
                try:
                    # Old variant: found = re.search(key+'(.*)'+keys_process[count+1],line)
                    found = re.search('(?<='+key+')(.*?)(?='+keys_process[count+1]+')', line).groups()[0]
                    sanitizedKey = key.rstrip(":")
                    found = found.lstrip(' ').rstrip(' ')
                    log[sanitizedKey]=found
                except:
                    pass

        if mode == "remote":
            bundle.append({"_index": "loki_logs", "body": log})
            if len(bundle) > 1000:
                print("INFO: Ingesting logs to the Elastic stack")
                shipLogs(esConn, bundle)
                bundle = []
        else: # Defaults to local mode
            bundle.append(log)

    if mode == "remote":
        print("INFO: Shipping remaining logs to the Elastic stack")
        shipLogs(esConn, bundle)
    else: # Defaults to local mode
        saveLogs(logfileName, bundle)


def configureElk(host, port):
    es = Elasticsearch(host=host, port=port)
    return es

def parseArgs():
    ap = argparse.ArgumentParser(description="Convert dull Loki logs to meaningful data and ingest to ElasticSearch!", epilog="Good, wasn't it?!")

    ap.add_argument("--file", metavar="File to import", required=True, help="Input the logfile to use for an import")
    ap.add_argument("--host", metavar="IP address of Elastic", required=False, default="localhost", help="Elastic host to use for a connection")
    ap.add_argument("--port", metavar="Port of Elastic", required=False, default=9200, help="Elastic port to use for a connection")
    ap.add_argument("--mode", metavar="Decide ingestion mode", required=False, default="local", choices=['local', 'remote'], help="Local parsing or a remote port of the logfile")

    args = vars(ap.parse_args())
    return args

def init():

    args = parseArgs()

    if args['mode'] == "local":
        esConn = "localhost" #TODO: See if this can be avoided
    else:
        esConn = configureElk(args['host'], args['port'])

    filePtr = open(args['file'])

    parseLogs(filePtr, esConn, args['mode'])

    print("INFO: Quitting. All logs have been parsed successfully.")



init()